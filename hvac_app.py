
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Electric HVAC-AC Entry Tool")
st.title("Electric HVAC-AC Form Filler")

st.markdown("""
This app allows you to upload a `.xlsm` workbook and a `.csv` file with multiple HVAC entries
that will populate the **Electric HVAC-AC** section (HVAC sheet).

Expected CSV columns:
- Location
- Unit Tag
- BTU/hr
- Tons
- SEER
- Proposed Efficiency
- Manufacturer
- Model
""")

uploaded_excel = st.file_uploader("Upload your macro-enabled Excel file (.xlsm)", type="xlsm")
uploaded_csv = st.file_uploader("Upload a CSV file with HVAC entries", type="csv")

if st.button("Process Entries"):
    if uploaded_excel and uploaded_csv:
        # Load Excel and CSV
        wb = load_workbook(filename=uploaded_excel, keep_vba=True)
        sheet = wb["HVAC"]
        df = pd.read_csv(uploaded_csv)

        # Column index mapping based on assumed layout
        column_map = {
            "Location": 2,
            "Unit Tag": 3,
            "BTU/hr": 13,
            "Tons": 12,
            "SEER": 14,
            "Proposed Efficiency": 15,
            "Manufacturer": 9,
            "Model": 10
        }

        # Start inserting after the header (assume from row 12)
        row = 12
        while sheet.cell(row=row, column=2).value not in (None, ""):
            row += 1

        # Insert each row from CSV
        for _, entry in df.iterrows():
            for key, col in column_map.items():
                sheet.cell(row=row, column=col, value=entry.get(key, ""))
            row += 1

        # Save to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Workbook updated successfully!")
        st.download_button(
            label="Download Updated Workbook",
            data=output,
            file_name="HVACTool_Updated.xlsm",
            mime="application/vnd.ms-excel.sheet.macroEnabled.12"
        )
    else:
        st.error("Please upload both a .xlsm file and a .csv file to proceed.")
